from fastapi import FastAPI, UploadFile, File, HTTPException
from fastapi.responses import FileResponse, HTMLResponse
from fastapi.middleware.cors import CORSMiddleware
import sqlite3, os, tempfile, re as _re
from datetime import datetime
from typing import Optional

app = FastAPI()
app.add_middleware(CORSMiddleware, allow_origins=["*"], allow_methods=["*"], allow_headers=["*"])

# Prevent browsers from caching API responses (stale cache caused 1-byte JSON bugs)
from starlette.middleware.base import BaseHTTPMiddleware
from starlette.requests import Request as _Req
class NoCacheAPIMiddleware(BaseHTTPMiddleware):
    async def dispatch(self, request: _Req, call_next):
        response = await call_next(request)
        if request.url.path.startswith("/api/"):
            response.headers["Cache-Control"] = "no-store, no-cache, must-revalidate"
            response.headers["Pragma"] = "no-cache"
        return response
app.add_middleware(NoCacheAPIMiddleware)

# ─── Telegram ─────────────────────────────────────────────────────────────────
TG_TOKEN = os.environ.get("TG_TOKEN", "")
TG_CHAT  = os.environ.get("TG_CHAT", "")

async def tg_send(text: str):
    import httpx
    url = f"https://api.telegram.org/bot{TG_TOKEN}/sendMessage"
    try:
        async with httpx.AsyncClient(timeout=10) as client:
            await client.post(url, json={"chat_id": TG_CHAT, "text": text, "parse_mode": "HTML"})
    except Exception as e:
        print(f"TG error: {e}")













DB_PATH = "/data/stocks.db"

def get_db():
    os.makedirs("/data", exist_ok=True)
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn

def init_db():
    conn = get_db()
    conn.execute("""CREATE TABLE IF NOT EXISTS stock_snapshots (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        date TEXT NOT NULL, sku_name TEXT NOT NULL,
        stock_qty REAL NOT NULL DEFAULT 0, uploaded_at TEXT NOT NULL,
        UNIQUE(date, sku_name))""")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_date ON stock_snapshots(date)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_sku ON stock_snapshots(sku_name)")
    conn.execute("""CREATE TABLE IF NOT EXISTS hidden_items (
        sku_base TEXT PRIMARY KEY,
        hidden_at TEXT NOT NULL)""")
    conn.execute("""CREATE TABLE IF NOT EXISTS sales_data (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        date TEXT NOT NULL, sku_name TEXT NOT NULL,
        qty REAL NOT NULL DEFAULT 0, revenue REAL NOT NULL DEFAULT 0,
        doc_type TEXT DEFAULT 'sale',
        UNIQUE(date, sku_name, doc_type))""")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_sales_date ON sales_data(date)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_sales_sku ON sales_data(sku_name)")
    conn.execute("""CREATE TABLE IF NOT EXISTS sku_costs (
        sku_base TEXT PRIMARY KEY,
        cost REAL NOT NULL DEFAULT 0,
        updated_at TEXT NOT NULL)""")
    conn.execute("""CREATE TABLE IF NOT EXISTS sku_adjustments (
        project_id TEXT NOT NULL DEFAULT '',
        sku_base TEXT NOT NULL,
        qty_adj INTEGER NOT NULL DEFAULT 0,
        updated_at TEXT NOT NULL,
        PRIMARY KEY (project_id, sku_base))""")
    conn.execute("""CREATE TABLE IF NOT EXISTS order_excluded (
        project_id TEXT NOT NULL DEFAULT '',
        sku_base TEXT NOT NULL,
        excluded_at TEXT NOT NULL,
        PRIMARY KEY (project_id, sku_base))""")
    conn.execute("""CREATE TABLE IF NOT EXISTS order_added (
        project_id TEXT NOT NULL DEFAULT '',
        sku_base TEXT NOT NULL,
        added_at TEXT NOT NULL,
        PRIMARY KEY (project_id, sku_base))""")
    conn.execute("""CREATE TABLE IF NOT EXISTS projects (
        id TEXT PRIMARY KEY,
        name TEXT NOT NULL,
        arrival_date TEXT NOT NULL,
        created_at TEXT NOT NULL)""")
    conn.execute("""CREATE TABLE IF NOT EXISTS transfers_snapshots (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        report_date TEXT NOT NULL,
        saved_at TEXT NOT NULL,
        warehouses TEXT NOT NULL,
        data TEXT NOT NULL,
        UNIQUE(report_date))""")
    conn.commit(); conn.close()

init_db()

def _strip_size(n):
    return _re.sub(r'[\s]*\([^)]*\)[\s]*$', '', str(n)).strip()

# Алиасы: все ключи → каноническое имя товара
# Майки переименованы в футболки; Lost in summer → Love in summer
SKU_ALIASES = {
    'Черная майка без рукавов "No plans"': 'Черная футболка без рукавов "No plans"',
    'Голубая майка без рукавов "Call me maybe never"': 'Голубая футболка без рукавов "Call me maybe never"',
    'Белая майка без рукавов "Boysmint"': 'Белая футболка без рукавов "Boysmint"',
    'Черная футболка без рукавов "Love in summer"': 'Черная футболка без рукавов "Lost in summer"',
}

def _canon_name(n):
    """Strip size suffix and apply canonical name aliases (майка→футболка, Lost→Love)."""
    base = _strip_size(n)
    return SKU_ALIASES.get(base, base)

_analytics_cache = None
_analytics_cache_key = None

def get_analytics_cache_key(conn):
    row = conn.execute("SELECT COUNT(*) as c, MAX(date) as d FROM stock_snapshots").fetchone()
    return f"{row['c']}_{row['d']}"

ANALYTICS_JSON_PATH = "/data/analytics_cache.json"

def build_analytics_data(conn):
    dates = [r[0] for r in conn.execute("SELECT DISTINCT date FROM stock_snapshots ORDER BY date").fetchall()]
    if not dates:
        return {"dates": [], "stock": {}}
    rows = conn.execute("""
        SELECT date, sku_name, SUM(stock_qty) as qty
        FROM stock_snapshots
        GROUP BY date, sku_name
        ORDER BY date, sku_name
    """).fetchall()
    stock = {}
    for r in rows:
        base = _canon_name(r["sku_name"])
        if base not in stock:
            stock[base] = {}
        stock[base][r["date"]] = stock[base].get(r["date"], 0) + r["qty"]
    return {"dates": dates, "stock": stock}

def build_turnover_data(analytics_data):
    """Pre-compute dis/cs/sea_days from analytics data. Called once on upload."""
    from datetime import date as _dt, timedelta
    dates = analytics_data.get("dates", [])
    stock = analytics_data.get("stock", {})
    if not dates:
        return {"dates": [], "skus": {}}
    cutoff = (_dt.today() - timedelta(days=365)).isoformat()
    recent = [d for d in dates if d >= cutoff]
    latest = dates[-1]
    def season(d):
        m = int(d[5:7])
        if m == 12 or m <= 2: return "winter"
        if m <= 5: return "spring"
        if m <= 8: return "summer"
        return "autumn"
    skus = {}
    for base, dm in stock.items():
        dis = 0; prev = 0
        sea = {"winter": 0, "spring": 0, "summer": 0, "autumn": 0}
        for d in recent:
            q = dm.get(d, prev)
            if q >= 3:
                dis += 1
                sea[season(d)] += 1
            prev = q
        skus[base] = {"dis": dis, "cs": int(dm.get(latest, 0)), "sea_days": sea}
    return {"dates": dates, "skus": skus}

def rebuild_analytics_json(conn):
    """Build analytics + turnover caches writing JSON to disk row-by-row (low memory)."""
    import json
    from datetime import date as _dt, timedelta

    dates = [r[0] for r in conn.execute(
        "SELECT DISTINCT date FROM stock_snapshots ORDER BY date").fetchall()]
    if not dates:
        empty = '{"dates":[],"stock":{}}'
        os.makedirs("/data", exist_ok=True)
        with open(ANALYTICS_JSON_PATH, "w") as f: f.write(empty)
        with open(TURNOVER_JSON_PATH, "w") as f: f.write('{"dates":[],"skus":{}}')
        return

    os.makedirs("/data", exist_ok=True)
    cutoff = (_dt.today() - timedelta(days=365)).isoformat()
    recent = [d for d in dates if d >= cutoff]
    latest = dates[-1]

    def season(d):
        m = int(d[5:7])
        if m == 12 or m <= 2: return "winter"
        if m <= 5: return "spring"
        if m <= 8: return "summer"
        return "autumn"

    # Stream analytics JSON to disk and compute turnover in one pass
    atmp = ANALYTICS_JSON_PATH + ".tmp"
    ttmp = TURNOVER_JSON_PATH + ".tmp"
    skus_turnover = {}  # compact: {base: {dis,cs,sea_days}}
    dates_json = json.dumps(dates, ensure_ascii=False)

    # Load sales data for all SKUs
    sales_rows = conn.execute(
        "SELECT sku_name, SUM(CASE WHEN doc_type='sale' THEN qty ELSE -qty END) as nq, "
        "SUM(CASE WHEN doc_type='sale' THEN revenue ELSE -revenue END) as nr, "
        "SUM(CASE WHEN doc_type='sale' THEN revenue ELSE -revenue END) / NULLIF(SUM(CASE WHEN doc_type='sale' THEN qty ELSE -qty END),0) as ap "
        "FROM sales_data GROUP BY sku_name"
    ).fetchall()
    sales_by_sku = {r["sku_name"]: {"nq": r["nq"] or 0, "nr": r["nr"] or 0, "ap": r["ap"] or 0} for r in sales_rows}

    # Load seasonal sales data
    sea_rows = conn.execute(
        "SELECT sku_name, "
        "SUM(CASE WHEN CAST(substr(date,6,2) AS INT) IN (12,1,2) THEN (CASE WHEN doc_type='sale' THEN revenue ELSE -revenue END) ELSE 0 END) as winter, "
        "SUM(CASE WHEN CAST(substr(date,6,2) AS INT) BETWEEN 3 AND 5 THEN (CASE WHEN doc_type='sale' THEN revenue ELSE -revenue END) ELSE 0 END) as spring, "
        "SUM(CASE WHEN CAST(substr(date,6,2) AS INT) BETWEEN 6 AND 8 THEN (CASE WHEN doc_type='sale' THEN revenue ELSE -revenue END) ELSE 0 END) as summer, "
        "SUM(CASE WHEN CAST(substr(date,6,2) AS INT) BETWEEN 9 AND 11 THEN (CASE WHEN doc_type='sale' THEN revenue ELSE -revenue END) ELSE 0 END) as autumn "
        "FROM sales_data GROUP BY sku_name"
    ).fetchall()
    sea_by_sku = {r["sku_name"]: {"winter": r["winter"] or 0, "spring": r["spring"] or 0,
                                   "summer": r["summer"] or 0, "autumn": r["autumn"] or 0} for r in sea_rows}

    # Build chart data (weekly revenue)
    chart_rows = conn.execute(
        "SELECT sku_name, date, SUM(CASE WHEN doc_type='sale' THEN revenue ELSE -revenue END) as rev "
        "FROM sales_data GROUP BY sku_name, date"
    ).fetchall()
    chart_by_sku = {}
    for r in chart_rows:
        chart_by_sku.setdefault(r["sku_name"], {})[r["date"]] = r["rev"]

    # Pre-aggregate sales by base name (stripping size suffixes)
    sales_by_base = {}
    sea_by_base = {}
    chart_by_base = {}
    for sku_name, s in sales_by_sku.items():
        base = _canon_name(sku_name)
        if base not in sales_by_base:
            sales_by_base[base] = {"nq": 0, "nr": 0}
        sales_by_base[base]["nq"] += s["nq"]
        sales_by_base[base]["nr"] += s["nr"]
    for base in sales_by_base:
        nq = sales_by_base[base]["nq"]
        nr = sales_by_base[base]["nr"]
        sales_by_base[base]["ap"] = nr / nq if nq > 0 else 0

    for sku_name, sea in sea_by_sku.items():
        base = _canon_name(sku_name)
        if base not in sea_by_base:
            sea_by_base[base] = {"winter":0,"spring":0,"summer":0,"autumn":0}
        for s in ("winter","spring","summer","autumn"):
            sea_by_base[base][s] += sea[s]

    for sku_name, chart_map in chart_by_sku.items():
        base = _canon_name(sku_name)
        if base not in chart_by_base:
            chart_by_base[base] = {}
        for d, v in chart_map.items():
            chart_by_base[base][d] = chart_by_base[base].get(d, 0) + v

    def get_sales(base_name):
        """Aggregate sales for a base SKU."""
        s = sales_by_base.get(base_name, {"nq":0,"nr":0,"ap":0})
        sea = sea_by_base.get(base_name, {"winter":0,"spring":0,"summer":0,"autumn":0})
        chart_map = chart_by_base.get(base_name, {})
        chart = [chart_map.get(d, 0) for d in dates]
        return s["nq"], s["nr"], s["ap"], sea, chart

    # Get all SKU names first (to iterate base by base)
    sku_names = [r[0] for r in conn.execute(
        "SELECT DISTINCT sku_name FROM stock_snapshots ORDER BY sku_name").fetchall()]

    # First pass: collect all per-SKU data + build base aggregates
    sku_data = {}   # sku_name -> {date: qty}
    base_agg = {}   # base -> {date: qty}
    for sku_name in sku_names:
        rows = conn.execute(
            "SELECT date, SUM(stock_qty) as qty FROM stock_snapshots "
            "WHERE sku_name=? GROUP BY date", (sku_name,)).fetchall()
        dm = {r["date"]: r["qty"] for r in rows}
        sku_data[sku_name] = dm
        base = _canon_name(sku_name)
        if base not in base_agg:
            base_agg[base] = {}
        for d, q in dm.items():
            base_agg[base][d] = base_agg[base].get(d, 0) + q

    with open(atmp, "w", encoding="utf-8") as af:
        af.write('{"dates":'); af.write(dates_json); af.write(',"stock":{')
        first = True
        # Write base-aggregate keys first (sum of all sizes)
        for base, dm in base_agg.items():
            if not first: af.write(",")
            first = False
            af.write(json.dumps(base, ensure_ascii=False))
            af.write(":")
            af.write(json.dumps(dm, ensure_ascii=False, separators=(",", ":")))
            # Turnover stats on the aggregate — use ALL dates (sales cover all time)
            dis = 0; prev = 0
            sea = {"winter":0,"spring":0,"summer":0,"autumn":0}
            for d in dates:
                q = dm.get(d, prev)
                if q >= 3: dis += 1; sea[season(d)] += 1
                prev = q
            nq, nr, ap, sea_rev, chart = get_sales(base)
            skus_turnover[base] = {"dis": dis, "cs": int(dm.get(latest, 0)), "sea_days": sea,
                                    "nq": nq, "nr": nr, "ap": ap,
                                    "sea": sea_rev, "chart": chart}
        # Write per-size keys only where sku_name has a size suffix
        # Use canonical name (футболка instead of майка) for the output key
        written_sizes = {}  # canonical_sku -> aggregated dm (handles майка+футболка same size)
        for sku_name, dm in sku_data.items():
            base = _canon_name(sku_name)
            if _strip_size(sku_name) == sku_name:
                continue  # no size suffix — already included in base_agg (even if aliased)
            size_suffix = sku_name[len(_strip_size(sku_name)):]  # e.g. " (L)"
            canonical_sku = base + size_suffix  # e.g. "Черная футболка без рукавов ... (L)"
            if canonical_sku not in written_sizes:
                written_sizes[canonical_sku] = dict(dm)
            else:
                for d, q in dm.items():
                    written_sizes[canonical_sku][d] = written_sizes[canonical_sku].get(d, 0) + q

        for canonical_sku, dm in written_sizes.items():
            af.write(",")
            af.write(json.dumps(canonical_sku, ensure_ascii=False))
            af.write(":")
            af.write(json.dumps(dm, ensure_ascii=False, separators=(",", ":")))
            # Turnover stats for this size (aggregate across aliases)
            sz_s = {"nq": 0, "nr": 0, "ap": 0}
            sz_sea = {"winter": 0, "spring": 0, "summer": 0, "autumn": 0}
            sz_chart_map = {}
            # Pull stats from all sku_names that map to this canonical_sku
            base_of_canon = _strip_size(canonical_sku)
            size_suffix_of_canon = canonical_sku[len(base_of_canon):]
            for orig_sku in sku_data:
                if _canon_name(orig_sku) + orig_sku[len(_strip_size(orig_sku)):] == canonical_sku or orig_sku == canonical_sku:
                    s = sales_by_sku.get(orig_sku, {"nq": 0, "nr": 0, "ap": 0})
                    sz_s["nq"] += s["nq"]; sz_s["nr"] += s["nr"]
                    sea = sea_by_sku.get(orig_sku, {"winter": 0, "spring": 0, "summer": 0, "autumn": 0})
                    for k in sea: sz_sea[k] += sea[k]
                    for d, v in chart_by_sku.get(orig_sku, {}).items():
                        sz_chart_map[d] = sz_chart_map.get(d, 0) + v
            sz_s["ap"] = sz_s["nr"] / sz_s["nq"] if sz_s["nq"] > 0 else 0
            sz_chart = [sz_chart_map.get(d, 0) for d in dates]
            sz_dis = 0; sz_prev = 0
            for d in dates:
                q = dm.get(d, sz_prev)
                if q >= 3: sz_dis += 1
                sz_prev = q
            skus_turnover[canonical_sku] = {
                "dis": sz_dis, "cs": int(dm.get(latest, 0)),
                "nq": sz_s["nq"], "nr": sz_s["nr"], "ap": sz_s["ap"],
                "sea": sz_sea, "chart": sz_chart
            }
        af.write("}}")
    os.replace(atmp, ANALYTICS_JSON_PATH)

    # Write turnover cache
    with open(ttmp, "w", encoding="utf-8") as tf:
        tf.write('{"dates":'); tf.write(dates_json)
        tf.write(',"skus":'); tf.write(json.dumps(skus_turnover, ensure_ascii=False, separators=(",", ":")))
        tf.write("}")
    os.replace(ttmp, TURNOVER_JSON_PATH)
    print(f"Caches rebuilt: {len(dates)} dates, {len(sku_names)} SKUs")

_CP1252_MAP = {
    0x20AC:0x80,0x201A:0x82,0x0192:0x83,0x201E:0x84,0x2026:0x85,
    0x2020:0x86,0x2021:0x87,0x02C6:0x88,0x2030:0x89,0x0160:0x8A,
    0x2039:0x8B,0x0152:0x8C,0x017D:0x8E,0x2018:0x91,0x2019:0x92,
    0x201C:0x93,0x201D:0x94,0x2022:0x95,0x2013:0x96,0x2014:0x97,
    0x02DC:0x98,0x2122:0x99,0x0161:0x9A,0x203A:0x9B,0x0153:0x9C,
    0x017E:0x9E,0x0178:0x9F,
}

def _fix_xlrd_str(s):
    """Fix mojibake: if xlrd returned UTF-8 bytes mis-decoded as cp1252/latin-1,
    convert back to proper Unicode. Returns original string if not mojibake."""
    if not s:
        return s
    # If all chars are pure ASCII or already Cyrillic (U+0400+) — no fix needed
    if all(ord(c) < 0x80 or ord(c) >= 0x400 for c in s):
        return s
    try:
        b = bytearray()
        for c in s:
            n = ord(c)
            if n in _CP1252_MAP:
                b.append(_CP1252_MAP[n])
            elif n < 256:
                b.append(n)
            else:
                return s  # genuine Unicode char — not mojibake
        result = b.decode('utf-8')
        return result
    except (UnicodeDecodeError, ValueError):
        return s

def _try_parse_date_str(val, rxls):
    """Try to parse a date from a string. Returns YYYY-MM-DD or None."""
    # DD.MM.YYYY  e.g. 21.05.2026  ← МойСклад default, check first
    m = rxls.search(r'(\d{1,2})\.(\d{1,2})\.(\d{4})', val)
    if m:
        d, mo, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
        if 1 <= mo <= 12 and 1 <= d <= 31 and y >= 2000:
            return "{}-{:02d}-{:02d}".format(y, mo, d)
    # MM/DD/YYYY  e.g. 05/21/2026
    m = rxls.search(r'(\d{1,2})/(\d{1,2})/(\d{4})', val)
    if m:
        a, b, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
        if y >= 2000:
            if a > 12:
                return "{}-{:02d}-{:02d}".format(y, b, a)
            else:
                return "{}-{:02d}-{:02d}".format(y, a, b)
    # YYYY-MM-DD
    m = rxls.search(r'(\d{4})-(\d{2})-(\d{2})', val)
    if m and int(m.group(1)) >= 2000:
        return m.group(0)
    return None

def _xlrd_serial_to_date(value, datemode, rxls):
    """Convert xlrd float to YYYY-MM-DD if it looks like a valid date serial."""
    import xlrd as _xl
    try:
        t = _xl.xldate_as_tuple(value, datemode)
        if 2015 <= t[0] <= 2035:
            return "{:04d}-{:02d}-{:02d}".format(*t[:3])
    except Exception:
        pass
    return None

def parse_xls(file_path):
    import xlrd, re as _rxls
    book = xlrd.open_workbook(file_path, encoding_override='utf-8')
    sheet = book.sheet_by_index(0)
    header_row = None
    name_col = None
    stock_col = None

    # ── Two-pass date detection: prioritize 'на момент' (actual stock date)
    # over 'отчет создан' (export timestamp)
    date_na_moment = None    # highest priority — actual snapshot date
    date_otchet = None       # fallback — export date
    date_serial = None       # last resort — bare Excel serial

    for i in range(min(25, sheet.nrows)):
        row_vals = [sheet.cell(i, j) for j in range(sheet.ncols)]

        # ── Find header row ──────────────────────────────────────────────────
        if header_row is None:
            row_strs = [str(c.value).strip() for c in row_vals]
            if any('аименование' in v for v in row_strs):
                header_row = i
                for j, v in enumerate(row_strs):
                    if 'аименование' in v: name_col = j
                    if 'статок' in v and 'умм' not in v.lower(): stock_col = j

        # ── Scan for date labels ─────────────────────────────────────────────
        for j, cell in enumerate(row_vals):
            if cell.ctype not in (1, 2, 3):
                continue
            val = str(cell.value).strip()

            # Label cell: check what label it is
            val_lower = val.lower()
            is_na_moment = 'на момент' in val_lower
            is_otchet = 'отчет создан' in val_lower or 'отчёт создан' in val_lower
            is_date_label = is_na_moment or is_otchet or 'дата' in val_lower

            if is_date_label:
                # Gather candidates from same cell + next 1-2 cells
                candidates = [val]
                for delta in (1, 2):
                    if j + delta < sheet.ncols:
                        nc = sheet.cell(i, j + delta)
                        candidates.append(str(nc.value).strip())
                        if nc.ctype == 3:
                            d = _xlrd_serial_to_date(nc.value, book.datemode, _rxls)
                            if d: candidates.append(d)
                for cand in candidates:
                    d = _try_parse_date_str(cand, _rxls)
                    if d:
                        if is_na_moment and date_na_moment is None:
                            date_na_moment = d
                        elif not is_na_moment and date_otchet is None:
                            date_otchet = d
                        break

            # Bare Excel date cell (no label nearby)
            elif cell.ctype == 3 and date_serial is None:
                d = _xlrd_serial_to_date(cell.value, book.datemode, _rxls)
                if d: date_serial = d

            # Bare float that looks like a date serial
            elif cell.ctype == 2 and date_serial is None:
                if 42005 <= cell.value <= 47848:
                    d = _xlrd_serial_to_date(cell.value, book.datemode, _rxls)
                    if d: date_serial = d

    # Pick best date: на момент > отчет создан > serial > today
    report_date = date_na_moment or date_otchet or date_serial
    if report_date is None:
        from datetime import date as _date
        report_date = _date.today().strftime('%Y-%m-%d')

    if header_row is None or name_col is None or stock_col is None:
        raise ValueError("Не найдена таблица с остатками (нет колонки Наименование/Остаток)")

    rows = []
    for i in range(header_row + 1, sheet.nrows):
        name = _fix_xlrd_str(str(sheet.cell_value(i, name_col)).strip())
        if not name or name in ('nan', 'Наименование', 'None'): continue
        try:
            qty = float(sheet.cell_value(i, stock_col))
        except Exception:
            qty = 0.0
        if qty <= 0: continue  # skip zero/negative stock rows
        rows.append({'sku_name': name, 'stock_qty': qty})
    if not rows:
        raise ValueError("Таблица пустая")
    return report_date, rows

@app.post("/api/upload")
async def upload_stock(file: UploadFile = File(...)):
    if not file.filename.lower().endswith('.xls'):
        raise HTTPException(400, "Только файлы .xls из МоегоСклада")
    with tempfile.NamedTemporaryFile(suffix='.xls', delete=False) as tmp:
        tmp.write(await file.read()); tmp_path = tmp.name
    try:
        date_str, rows = parse_xls(tmp_path)
    except Exception as e:
        os.unlink(tmp_path); raise HTTPException(400, str(e))
    finally:
        if os.path.exists(tmp_path): os.unlink(tmp_path)
    conn = get_db()
    uploaded_at = datetime.now().isoformat()
    inserted = 0
    updated = 0
    # Log first SKU to diagnose encoding on server
    if rows:
        import logging
        logging.warning(f"UPLOAD first SKU repr: {repr(rows[0]['sku_name'])}")
        logging.warning(f"UPLOAD has nowhere: {any('owhere' in r['sku_name'] for r in rows)}")
    for row in rows:
        conn.execute(
            "INSERT OR REPLACE INTO stock_snapshots (date,sku_name,stock_qty,uploaded_at) VALUES (?,?,?,?)",
            (date_str, row['sku_name'], row['stock_qty'], uploaded_at))
        inserted += 1
    conn.commit(); conn.close()
    global _analytics_cache, _analytics_cache_key
    _analytics_cache = None
    _analytics_cache_key = None
    conn2 = get_db()
    rebuild_analytics_json(conn2)
    conn2.close()
    return {"date": date_str, "inserted": inserted, "total_skus": len(rows),
            "_debug_sample": [r['sku_name'] for r in rows if 'owhere' in r['sku_name'].lower()][:3]}


@app.post("/api/debug-parse-xls")
async def debug_parse_xls(file: UploadFile = File(...)):
    """Debug: show exactly what xlrd returns for SKU names, with repr()."""
    import xlrd
    with tempfile.NamedTemporaryFile(suffix='.xls', delete=False) as tmp:
        tmp.write(await file.read()); tmp_path = tmp.name
    try:
        results = {}
        for enc in [None, 'utf-8', 'cp1251', 'latin-1']:
            try:
                kw = {'encoding_override': enc} if enc else {}
                book = xlrd.open_workbook(tmp_path, **kw)
                sheet = book.sheet_by_index(0)
                names = []
                for i in range(min(sheet.nrows, 500)):
                    v = sheet.cell_value(i, 3)
                    if v and 'owhere' in str(v).lower():
                        names.append({'row': i, 'repr': repr(str(v)), 'value': str(v)})
                results[str(enc)] = {'found': len(names), 'samples': names[:3]}
            except Exception as e:
                results[str(enc)] = {'error': str(e)}
        return results
    finally:
        if os.path.exists(tmp_path): os.unlink(tmp_path)

@app.delete("/api/debug-delete-sku")
def debug_delete_sku(q: str):
    """Delete all stock_snapshots rows where sku_name contains q (for fixing corrupted data)."""
    conn = get_db()
    count = conn.execute(
        "SELECT COUNT(*) as c FROM stock_snapshots WHERE LOWER(sku_name) LIKE ?",
        (f"%{q.lower()}%",)
    ).fetchone()["c"]
    conn.execute("DELETE FROM stock_snapshots WHERE LOWER(sku_name) LIKE ?", (f"%{q.lower()}%",))
    conn.commit(); conn.close()
    global _analytics_cache, _analytics_cache_key
    _analytics_cache = None; _analytics_cache_key = None
    import os
    if os.path.exists(ANALYTICS_JSON_PATH): os.remove(ANALYTICS_JSON_PATH)
    if os.path.exists(TURNOVER_JSON_PATH): os.remove(TURNOVER_JSON_PATH)
    return {"deleted": count, "query": q}

@app.get("/api/debug-stock-totals")
def debug_stock_totals():
    """Show total stock per date to find dates with inflated numbers."""
    conn = get_db()
    rows = conn.execute(
        "SELECT date, COUNT(DISTINCT sku_name) as sku_count, SUM(stock_qty) as total_qty "
        "FROM stock_snapshots GROUP BY date ORDER BY date DESC LIMIT 30"
    ).fetchall()
    conn.close()
    return [{"date": r["date"], "sku_count": r["sku_count"], "total_qty": r["total_qty"]} for r in rows]

@app.get("/api/debug-sku-search")
def debug_sku_search(q: str = "nowhere"):
    """Search for SKU names in DB containing the query string (case-insensitive)."""
    conn = get_db()
    rows = conn.execute(
        "SELECT DISTINCT sku_name, COUNT(*) as dates, SUM(stock_qty) as total_qty "
        "FROM stock_snapshots WHERE LOWER(sku_name) LIKE ? "
        "GROUP BY sku_name ORDER BY sku_name",
        (f"%{q.lower()}%",)
    ).fetchall()
    conn.close()
    results = [{"sku_name": r["sku_name"], "repr": repr(r["sku_name"]),
                "dates": r["dates"], "total_qty": r["total_qty"],
                "stripped": _re.sub(r'[\s]*\([^)]*\)[\s]*$', '', str(r["sku_name"])).strip()}
               for r in rows]
    return {"count": len(results), "results": results}

@app.get("/api/debug-skus")
def debug_skus():
    """Show sample sku_names to understand naming format."""
    conn = get_db()
    rows = conn.execute(
        "SELECT DISTINCT sku_name FROM stock_snapshots ORDER BY sku_name LIMIT 200"
    ).fetchall()
    conn.close()
    names = [r["sku_name"] for r in rows]
    has_parens = [n for n in names if "(" in n]
    has_slash   = [n for n in names if "/" in n]
    return {
        "total": len(names),
        "sample": names[:40],
        "has_parens": has_parens[:30],
        "has_slash": has_slash[:20],
    }

@app.get("/api/dates")
def get_dates():
    conn = get_db()
    rows = conn.execute("SELECT date, COUNT(DISTINCT sku_name) as sku_count FROM stock_snapshots GROUP BY date ORDER BY date DESC").fetchall()
    conn.close()
    return [{"date": r["date"], "sku_count": r["sku_count"]} for r in rows]

@app.get("/api/stocks")
def get_stocks(date: Optional[str]=None, search: Optional[str]=None, page: int=1, per_page: int=50):
    conn = get_db()
    if not date:
        row = conn.execute("SELECT MAX(date) as d FROM stock_snapshots").fetchone()
        date = row["d"]
    if not date:
        return {"date": None, "items": [], "total": 0, "pages": 0}
    cond = ["date = ?"]; params = [date]
    if search:
        cond.append("LOWER(sku_name) LIKE ?"); params.append(f"%{search.lower()}%")
    where = " AND ".join(cond)
    total = conn.execute(f"SELECT COUNT(*) as c FROM stock_snapshots WHERE {where}", params).fetchone()["c"]
    rows = conn.execute(f"SELECT sku_name, stock_qty FROM stock_snapshots WHERE {where} ORDER BY sku_name LIMIT ? OFFSET ?",
                        params+[per_page,(page-1)*per_page]).fetchall()
    conn.close()
    return {"date": date, "items": [{"sku_name": r["sku_name"], "stock_qty": r["stock_qty"]} for r in rows],
            "total": total, "pages": -(-total//per_page)}

@app.get("/api/stocks/all")
def get_all_stocks():
    """Serve pre-built analytics cache from disk. Zero memory usage."""
    if os.path.exists(ANALYTICS_JSON_PATH) and os.path.getsize(ANALYTICS_JSON_PATH) > 100:
        return FileResponse(ANALYTICS_JSON_PATH, media_type="application/json")
    # Cache missing — build it
    conn = get_db()
    rebuild_analytics_json(conn)
    conn.close()
    if os.path.exists(ANALYTICS_JSON_PATH):
        return FileResponse(ANALYTICS_JSON_PATH, media_type="application/json")
    return {"dates": [], "stock": {}}

@app.get("/api/stats")
def get_stats():
    conn = get_db()
    r1 = conn.execute("SELECT COUNT(*) as c FROM stock_snapshots").fetchone()["c"]
    r2 = conn.execute("SELECT COUNT(DISTINCT sku_name) as c FROM stock_snapshots").fetchone()["c"]
    r3 = conn.execute("SELECT COUNT(DISTINCT date) as c FROM stock_snapshots").fetchone()["c"]
    dr = conn.execute("SELECT MIN(date) as mn, MAX(date) as mx FROM stock_snapshots").fetchone()
    conn.close()
    return {"total_records": r1, "total_skus": r2, "total_dates": r3, "date_from": dr["mn"], "date_to": dr["mx"]}

@app.post("/api/hide")
async def hide_item(data: dict):
    conn = get_db()
    conn.execute("INSERT OR IGNORE INTO hidden_items (sku_base, hidden_at) VALUES (?,?)",
                 (data["sku_base"], datetime.now().isoformat()))
    conn.commit(); conn.close()
    return {"ok": True}

@app.delete("/api/hide/{sku_base}")
def unhide_item(sku_base: str):
    conn = get_db()
    conn.execute("DELETE FROM hidden_items WHERE sku_base=?", (sku_base,))
    conn.commit(); conn.close()
    return {"ok": True}

@app.get("/api/hidden")
def get_hidden():
    conn = get_db()
    rows = conn.execute("SELECT sku_base FROM hidden_items ORDER BY hidden_at DESC").fetchall()
    conn.close()
    return [r["sku_base"] for r in rows]

@app.post("/api/upload-sales")
async def upload_sales(file: UploadFile = File(...)):
    import csv, io
    if not file.filename.lower().endswith(".csv"):
        raise HTTPException(400, "Только CSV файлы")
    content_bytes = await file.read()
    try:
        text = content_bytes.decode("utf-8-sig")
    except:
        text = content_bytes.decode("cp1251")

    reader = csv.DictReader(io.StringIO(text))
    rows = list(reader)

    if not rows:
        raise HTTPException(400, "Файл пустой")

    is_return = "возврат" in file.filename.lower() or "возврат" in (rows[0].get("Документ","")).lower()
    doc_type = "return" if is_return else "sale"

    import pandas as pd
    df = pd.DataFrame(rows)

    if "Артикул" in df.columns:
        df = df[df["Артикул"].notna() & (df["Артикул"].astype(str).str.strip() != "")]

    df["Дата"] = pd.to_datetime(df["Дата документа"], dayfirst=True).dt.date
    for col in ["Количество", "Сумма"]:
        df[col] = (df[col].astype(str)
                   .str.replace("\xa0", "", regex=False)
                   .str.replace(" ", "", regex=False)
                   .str.replace(",", ".", regex=False))
        df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)

    conn = get_db()
    from collections import defaultdict
    agg = defaultdict(lambda: {"qty": 0.0, "revenue": 0.0})
    for _, row in df.iterrows():
        key = (str(row["Дата"]), str(row["Наименование"]).strip())
        agg[key]["qty"]     += float(row["Количество"])
        agg[key]["revenue"] += float(row["Сумма"])
    dates = list({k[0] for k in agg.keys()})
    placeholders = ",".join("?" * len(dates))
    conn.execute(f"DELETE FROM sales_data WHERE doc_type=? AND date IN ({placeholders})", [doc_type] + dates)
    inserted = 0
    for (date, sku_name), vals in agg.items():
        try:
            conn.execute(
                "INSERT OR REPLACE INTO sales_data (date, sku_name, qty, revenue, doc_type) VALUES (?,?,?,?,?)",
                (date, sku_name, vals["qty"], vals["revenue"], doc_type)
            )
            inserted += 1
        except: pass
    conn.commit()
    date_from = str(df["Дата"].min())
    date_to = str(df["Дата"].max())
    conn.close()
    return {"inserted": inserted, "doc_type": doc_type, "date_from": date_from, "date_to": date_to}

@app.post("/api/check-bestsellers")
async def check_bestsellers():
    """Проверяет бестселлеры (turn >= 2000) у которых запас < 45 дней и шлёт пуш в Telegram."""
    import httpx
    conn = get_db()
    # Берём последние остатки по каждому SKU
    rows = conn.execute("""
        SELECT sku_name, stock_qty
        FROM stock_snapshots
        WHERE date = (SELECT MAX(date) FROM stock_snapshots)
    """).fetchall()
    # Берём продажи за последние 90 дней для расчёта дневных продаж
    sales = conn.execute("""
        SELECT sku_name, SUM(CASE WHEN doc_type='sale' THEN qty ELSE -qty END) as total_qty
        FROM sales_data
        WHERE date >= date('now', '-90 days')
        GROUP BY sku_name
    """).fetchall()
    conn.close()

    sales_map = {_canon_name(r["sku_name"]): r["total_qty"] for r in sales}

    alerts = []
    for r in rows:
        base = _canon_name(r["sku_name"])
        stock = r["stock_qty"]
        if stock <= 0:
            continue
        qty_90 = sales_map.get(base, 0)
        if qty_90 <= 0:
            continue
        daily = qty_90 / 90
        turn_rub = daily  # используем штуки/день для сравнения
        days_left = stock / daily
        # Только бестселлеры (продаётся хотя бы 1 шт в 2 дня) у которых < 45 дней запаса
        if daily >= 0.5 and days_left < 45:
            order_90 = max(0, round(daily * 90 - stock))
            alerts.append({
                "name": base,
                "stock": round(stock),
                "days_left": round(days_left),
                "daily": round(daily, 1),
                "order_90": order_90,
            })

    if not alerts:
        return {"sent": 0, "message": "Всё в порядке, критических остатков нет"}

    alerts.sort(key=lambda x: x["days_left"])
    lines = ["🚨 <b>Заканчиваются бестселлеры!</b>\n"]
    for a in alerts:
        lines.append(
            f"📦 <b>{a['name']}</b>\n"
            f"   Остаток: {a['stock']} шт · закончится через <b>{a['days_left']} дн.</b>\n"
            f"   К заказу на 90 дней: <b>{a['order_90']} шт</b>\n"
        )
    await tg_send("\n".join(lines))
    return {"sent": len(alerts), "alerts": alerts}


@app.get("/order")
def serve_order():
    if os.path.exists("order.html"):
        return FileResponse("order.html", media_type="text/html")
    return FileResponse("index.html", media_type="text/html")

@app.get("/turnover")
def serve_turnover():
    if os.path.exists("turnover.html"):
        return FileResponse("turnover.html", media_type="text/html")
    return FileResponse("index.html", media_type="text/html")

@app.get("/api/analytics-data")
def get_analytics_data():
    """Serve pre-built analytics JSON from disk. If missing or empty, build it first."""
    if os.path.exists(ANALYTICS_JSON_PATH) and os.path.getsize(ANALYTICS_JSON_PATH) > 100:
        return FileResponse(ANALYTICS_JSON_PATH, media_type="application/json")
    # File missing or empty — delete and rebuild
    if os.path.exists(ANALYTICS_JSON_PATH):
        os.remove(ANALYTICS_JSON_PATH)
    # First run: build the file
    conn = get_db()
    rebuild_analytics_json(conn)
    conn.close()
    if os.path.exists(ANALYTICS_JSON_PATH):
        return FileResponse(ANALYTICS_JSON_PATH, media_type="application/json")
    # Fallback: return empty
    return {"dates": [], "stock": {}}

@app.post("/api/rebuild-analytics")
def trigger_rebuild():
    """Manually trigger analytics rebuild (admin use)."""
    conn = get_db()
    rebuild_analytics_json(conn)
    conn.close()
    return {"ok": True}

@app.post("/api/invalidate-cache")
def invalidate_cache():
    global _analytics_cache, _analytics_cache_key
    _analytics_cache = None
    _analytics_cache_key = None
    # Also remove disk cache so next request rebuilds it
    if os.path.exists(ANALYTICS_JSON_PATH):
        os.remove(ANALYTICS_JSON_PATH)
    return {"ok": True}

@app.get("/analytics")
def serve_analytics():
    if os.path.exists("analytics.html"):
        return FileResponse("analytics.html", media_type="text/html")
    return FileResponse("index.html", media_type="text/html")


# ─── Costs ────────────────────────────────────────────────────────────────────
@app.get("/api/costs")
def get_costs():
    conn = get_db()
    rows = conn.execute("SELECT sku_base, cost FROM sku_costs").fetchall()
    conn.close()
    return {r["sku_base"]: r["cost"] for r in rows}

@app.post("/api/costs")
async def set_cost(data: dict):
    conn = get_db()
    conn.execute("INSERT OR REPLACE INTO sku_costs (sku_base, cost, updated_at) VALUES (?,?,?)",
                 (data["sku_base"], float(data.get("cost", 0)), datetime.now().isoformat()))
    conn.commit(); conn.close()
    return {"ok": True}

# ─── Adjustments ──────────────────────────────────────────────────────────────
@app.get("/api/adjustments")
def get_adjustments(project_id: str = ""):
    conn = get_db()
    rows = conn.execute("SELECT sku_base, qty_adj FROM sku_adjustments WHERE project_id=?", (project_id,)).fetchall()
    conn.close()
    return {r["sku_base"]: r["qty_adj"] for r in rows}

@app.post("/api/adjustments")
async def set_adjustment(data: dict):
    conn = get_db()
    adj = int(data.get("qty_adj", 0))
    pid = data.get("project_id", "")
    if adj == 0:
        conn.execute("DELETE FROM sku_adjustments WHERE project_id=? AND sku_base=?", (pid, data["sku_base"]))
    else:
        conn.execute("INSERT OR REPLACE INTO sku_adjustments (project_id, sku_base, qty_adj, updated_at) VALUES (?,?,?,?)",
                     (pid, data["sku_base"], adj, datetime.now().isoformat()))
    conn.commit(); conn.close()
    return {"ok": True}

# ─── Order excluded ───────────────────────────────────────────────────────────
@app.get("/api/excluded")
def get_excluded(project_id: str = ""):
    conn = get_db()
    rows = conn.execute("SELECT sku_base FROM order_excluded WHERE project_id=?", (project_id,)).fetchall()
    conn.close()
    return [r["sku_base"] for r in rows]

@app.post("/api/excluded")
async def add_excluded(data: dict):
    conn = get_db()
    pid = data.get("project_id", "")
    conn.execute("INSERT OR IGNORE INTO order_excluded (project_id, sku_base, excluded_at) VALUES (?,?,?)",
                 (pid, data["sku_base"], datetime.now().isoformat()))
    conn.commit(); conn.close()
    return {"ok": True}

@app.delete("/api/excluded/{sku_base}")
def remove_excluded(sku_base: str, project_id: str = ""):
    conn = get_db()
    conn.execute("DELETE FROM order_excluded WHERE project_id=? AND sku_base=?", (project_id, sku_base))
    conn.commit(); conn.close()
    return {"ok": True}

# ─── Order added (possible→active) ───────────────────────────────────────────
@app.get("/api/order-added")
def get_order_added(project_id: str = ""):
    conn = get_db()
    rows = conn.execute("SELECT sku_base FROM order_added WHERE project_id=?", (project_id,)).fetchall()
    conn.close()
    return [r["sku_base"] for r in rows]

@app.post("/api/order-added")
async def add_order_added(data: dict):
    conn = get_db()
    pid = data.get("project_id", "")
    conn.execute("INSERT OR IGNORE INTO order_added (project_id, sku_base, added_at) VALUES (?,?,?)",
                 (pid, data["sku_base"], datetime.now().isoformat()))
    conn.commit(); conn.close()
    return {"ok": True}

# ─── Projects ─────────────────────────────────────────────────────────────────
@app.get("/api/projects/{project_id}/data")
def get_project_data(project_id: str):
    """Load saved project state: adjustments, excluded, added."""
    conn = get_db()
    adj = {r["sku_base"]: r["qty_adj"] for r in conn.execute(
        "SELECT sku_base, qty_adj FROM sku_adjustments WHERE project_id=?", (project_id,)).fetchall()}
    excl = [r["sku_base"] for r in conn.execute(
        "SELECT sku_base FROM order_excluded WHERE project_id=?", (project_id,)).fetchall()]
    added = [r["sku_base"] for r in conn.execute(
        "SELECT sku_base FROM order_added WHERE project_id=?", (project_id,)).fetchall()]
    conn.close()
    return {"adjustments": adj, "excluded": excl, "added": added}

@app.post("/api/projects/{project_id}/data")
async def save_project_data(project_id: str, data: dict):
    """Save full project state at once: adjustments, excluded, added."""
    conn = get_db()
    now = datetime.now().isoformat()
    # Replace adjustments
    conn.execute("DELETE FROM sku_adjustments WHERE project_id=?", (project_id,))
    for base, qty in (data.get("adjustments") or {}).items():
        if int(qty) != 0:
            conn.execute("INSERT INTO sku_adjustments (project_id,sku_base,qty_adj,updated_at) VALUES (?,?,?,?)",
                         (project_id, base, int(qty), now))
    # Replace excluded
    conn.execute("DELETE FROM order_excluded WHERE project_id=?", (project_id,))
    for base in (data.get("excluded") or []):
        conn.execute("INSERT INTO order_excluded (project_id,sku_base,excluded_at) VALUES (?,?,?)",
                     (project_id, base, now))
    # Replace added
    conn.execute("DELETE FROM order_added WHERE project_id=?", (project_id,))
    for base in (data.get("added") or []):
        conn.execute("INSERT INTO order_added (project_id,sku_base,added_at) VALUES (?,?,?)",
                     (project_id, base, now))
    conn.commit(); conn.close()
    return {"ok": True}

@app.get("/api/projects")
def list_projects():
    conn = get_db()
    rows = conn.execute("SELECT id, name, arrival_date, created_at FROM projects ORDER BY created_at DESC").fetchall()
    conn.close()
    return [{"id": r["id"], "name": r["name"], "date": r["arrival_date"]} for r in rows]

@app.post("/api/projects")
async def upsert_project(data: dict):
    import uuid as _uuid
    conn = get_db()
    pid = data.get("id") or _uuid.uuid4().hex[:8]
    conn.execute("INSERT OR REPLACE INTO projects (id, name, arrival_date, created_at) VALUES (?,?,?,?)",
                 (pid, data["name"], data["date"], datetime.now().isoformat()))
    conn.commit(); conn.close()
    return {"id": pid, "name": data["name"], "date": data["date"]}

@app.delete("/api/projects/{project_id}")
def delete_project(project_id: str):
    conn = get_db()
    for tbl in ("sku_adjustments", "order_excluded", "order_added"):
        conn.execute(f"DELETE FROM {tbl} WHERE project_id=?", (project_id,))
    conn.execute("DELETE FROM projects WHERE id=?", (project_id,))
    conn.commit(); conn.close()
    return {"ok": True}


# ─── Turnover data: pre-computed per-SKU stats (fast) ────────────────────────
TURNOVER_JSON_PATH = "/data/turnover_cache.json"

@app.get("/api/turnover-data")
def get_turnover_data():
    """Serve pre-computed compact turnover stats from disk."""
    if os.path.exists(TURNOVER_JSON_PATH) and os.path.getsize(TURNOVER_JSON_PATH) > 100:
        # Validate cache has dis field — if not, invalidate and rebuild
        try:
            import json as _j
            with open(TURNOVER_JSON_PATH, "r", encoding="utf-8") as _f:
                _sample = _j.load(_f)
            _skus = _sample.get("skus", {})
            _first = next(iter(_skus.values()), {}) if _skus else {}
            if "dis" not in _first:
                os.remove(TURNOVER_JSON_PATH)
            else:
                return FileResponse(TURNOVER_JSON_PATH, media_type="application/json")
        except Exception:
            pass
    # Turnover cache missing — build from analytics cache if it exists (no SQL/memory)
    if os.path.exists(ANALYTICS_JSON_PATH):
        import json
        with open(ANALYTICS_JSON_PATH, "r", encoding="utf-8") as f:
            data = json.load(f)
        tdata = build_turnover_data(data)
        del data
        os.makedirs("/data", exist_ok=True)
        tmp = TURNOVER_JSON_PATH + ".tmp"
        with open(tmp, "w", encoding="utf-8") as f:
            json.dump(tdata, f, ensure_ascii=False, separators=(",", ":"))
        os.replace(tmp, TURNOVER_JSON_PATH)
        del tdata
        return FileResponse(TURNOVER_JSON_PATH, media_type="application/json")
    # Neither cache exists — need fresh rebuild from DB
    conn = get_db()
    rebuild_analytics_json(conn)
    conn.close()
    if os.path.exists(TURNOVER_JSON_PATH):
        return FileResponse(TURNOVER_JSON_PATH, media_type="application/json")
    return {"dates": [], "skus": {}}

@app.get("/sales")
def serve_sales():
    if os.path.exists("sales.html"):
        return FileResponse("sales.html", media_type="text/html")
    return FileResponse("index.html", media_type="text/html")

@app.get("/api/sales-monthly")
def get_sales_monthly(month: Optional[str] = None):
    """Monthly sales by canonical SKU with size breakdown, sorted by net revenue."""
    conn = get_db()
    # Available months
    months = [r[0] for r in conn.execute(
        "SELECT DISTINCT substr(date,1,7) as m FROM sales_data ORDER BY m DESC"
    ).fetchall()]
    if not months:
        conn.close()
        return {"months": [], "month": None,
                "summary": {"sales": 0, "returns": 0, "net": 0, "sales_qty": 0, "returns_qty": 0},
                "items": []}
    if not month or month not in months:
        month = months[0]
    rows = conn.execute(
        "SELECT sku_name, doc_type, SUM(qty) as qty, SUM(revenue) as rev "
        "FROM sales_data WHERE substr(date,1,7)=? GROUP BY sku_name, doc_type",
        (month,)
    ).fetchall()
    conn.close()

    bases = {}
    for r in rows:
        sku = r["sku_name"]
        base = _canon_name(sku)
        is_sale = r["doc_type"] == "sale"
        qty = r["qty"] or 0
        rev = r["rev"] or 0
        if base not in bases:
            bases[base] = {"sale_rev": 0, "sale_qty": 0, "ret_rev": 0, "ret_qty": 0, "sizes": {}}
        if is_sale:
            bases[base]["sale_rev"] += rev; bases[base]["sale_qty"] += qty
        else:
            bases[base]["ret_rev"] += rev; bases[base]["ret_qty"] += qty
        # size handling
        stripped = _strip_size(sku)
        if stripped != sku:
            size_raw = sku[len(stripped):].strip()   # e.g. "(L)"
            canonical_base = _canon_name(sku)
            size_key = size_raw
            if size_key not in bases[canonical_base]["sizes"]:
                bases[canonical_base]["sizes"][size_key] = {"sale_rev": 0, "sale_qty": 0, "ret_rev": 0, "ret_qty": 0}
            sz = bases[canonical_base]["sizes"][size_key]
            if is_sale: sz["sale_rev"] += rev; sz["sale_qty"] += qty
            else:       sz["ret_rev"]  += rev; sz["ret_qty"]  += qty

    total_sales = sum(d["sale_rev"] for d in bases.values())
    total_returns = sum(d["ret_rev"] for d in bases.values())
    total_sales_qty = sum(d["sale_qty"] for d in bases.values())
    total_returns_qty = sum(d["ret_qty"] for d in bases.values())

    items = []
    for base, d in bases.items():
        net = d["sale_rev"] - d["ret_rev"]
        sizes = sorted([
            {"name": sk, "sale_rev": sv["sale_rev"], "sale_qty": sv["sale_qty"],
             "ret_rev": sv["ret_rev"], "ret_qty": sv["ret_qty"],
             "net": sv["sale_rev"] - sv["ret_rev"]}
            for sk, sv in d["sizes"].items()
        ], key=lambda x: x["net"], reverse=True)
        items.append({
            "base": base, "sale_rev": d["sale_rev"], "sale_qty": d["sale_qty"],
            "ret_rev": d["ret_rev"], "ret_qty": d["ret_qty"], "net": net, "sizes": sizes
        })
    items.sort(key=lambda x: x["net"], reverse=True)

    return {
        "months": months, "month": month,
        "summary": {"sales": total_sales, "returns": total_returns,
                    "net": total_sales - total_returns,
                    "sales_qty": total_sales_qty, "returns_qty": total_returns_qty},
        "items": items
    }

@app.get("/forecast")
def serve_forecast():
    if os.path.exists("forecast.html"):
        return FileResponse("forecast.html", media_type="text/html")
    return FileResponse("index.html", media_type="text/html")

@app.get("/transfers")
def serve_transfers():
    if os.path.exists("transfers.html"):
        return FileResponse("transfers.html", media_type="text/html")
    return FileResponse("index.html", media_type="text/html")

@app.get("/revenue")
def serve_revenue():
    if os.path.exists("revenue.html"):
        return FileResponse("revenue.html", media_type="text/html")
    return FileResponse("index.html", media_type="text/html")

@app.get("/api/revenue-data")
def get_revenue_data():
    """Gross sales and returns per SKU base name, with per-size breakdown."""
    conn = get_db()
    rows = conn.execute("""
        SELECT sku_name,
            SUM(CASE WHEN doc_type='sale'   THEN qty     ELSE 0 END) as sale_qty,
            SUM(CASE WHEN doc_type='sale'   THEN revenue ELSE 0 END) as sale_rev,
            SUM(CASE WHEN doc_type='return' THEN qty     ELSE 0 END) as ret_qty,
            SUM(CASE WHEN doc_type='return' THEN revenue ELSE 0 END) as ret_rev
        FROM sales_data
        GROUP BY sku_name
    """).fetchall()
    conn.close()
    skus = {}
    for r in rows:
        base = _canon_name(r["sku_name"])
        if base not in skus:
            skus[base] = {"sale_qty": 0, "sale_rev": 0, "ret_qty": 0, "ret_rev": 0, "sizes": {}}
        skus[base]["sale_qty"] += r["sale_qty"] or 0
        skus[base]["sale_rev"] += r["sale_rev"] or 0
        skus[base]["ret_qty"]  += r["ret_qty"]  or 0
        skus[base]["ret_rev"]  += r["ret_rev"]  or 0
        if r["sku_name"] != base:
            skus[base]["sizes"][r["sku_name"]] = {
                "sale_qty": r["sale_qty"] or 0,
                "sale_rev": r["sale_rev"] or 0,
                "ret_qty":  r["ret_qty"]  or 0,
                "ret_rev":  r["ret_rev"]  or 0,
            }
    return skus

@app.post("/api/parse-stock")
async def parse_stock_file(file: UploadFile = File(...)):
    import csv, io, tempfile as _tmp
    raw = await file.read()
    ext = (file.filename or "").rsplit(".", 1)[-1].lower()
    warehouse = None
    report_date = None
    items = {}
    if ext in ("xls", "xlsx"):
        with _tmp.NamedTemporaryFile(suffix="." + ext, delete=False) as tf:
            tf.write(raw); tmp_path = tf.name
        try:
            import xlrd
            book = xlrd.open_workbook(tmp_path)
            sheet = book.sheet_by_index(0)
            # Extract warehouse name
            for i in range(min(20, sheet.nrows)):
                for j in range(sheet.ncols):
                    if str(sheet.cell_value(i, j)).strip() == "склад:" and j+1 < sheet.ncols:
                        warehouse = str(sheet.cell_value(i, j+1)).strip()
            # Use robust parse_xls for date + rows
            report_date, rows = parse_xls(tmp_path)
            for r in rows:
                if r["stock_qty"] > 0:
                    items[r["sku_name"]] = items.get(r["sku_name"], 0) + int(r["stock_qty"])
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Ошибка чтения XLS: {e}")
        finally:
            if os.path.exists(tmp_path): os.unlink(tmp_path)
    else:
        text = None
        for enc in ("utf-8", "cp1251", "latin-1"):
            try: text = raw.decode(enc); break
            except: continue
        if not text:
            raise HTTPException(status_code=400, detail="Не удалось декодировать CSV")
        import re as _re2
        all_rows = list(csv.reader(io.StringIO(text)))
        for row in all_rows:
            for i, cell in enumerate(row):
                if cell.strip() == "склад:" and i+1 < len(row) and not warehouse:
                    warehouse = row[i+1].strip()
                if "отчет создан" in cell.lower() and i+1 < len(row) and not report_date:
                    m = _re2.search(r"(\d{1,2})/(\d{1,2})/(\d{4})", row[i+1])
                    if m: report_date = f"{m.group(3)}-{int(m.group(1)):02d}-{int(m.group(2)):02d}"
        hdr = nc = qc = None
        for i, row in enumerate(all_rows):
            if "Наименование" in row:
                hdr = i; nc = row.index("Наименование")
                qc = next((j for j,c in enumerate(row) if c.strip()=="Остаток"), None); break
        if hdr is None: raise HTTPException(400, "Нет колонки Наименование")
        for row in all_rows[hdr+1:]:
            if len(row) <= nc: continue
            name = row[nc].strip()
            if not name: continue
            if qc is None or qc >= len(row): continue
            try: qty = float(row[qc].replace(",",".").replace("\xa0",""))
            except: continue
            if qty > 0: items[name] = items.get(name, 0) + int(qty)
    return {"warehouse": warehouse or "Склад", "items": items, "report_date": report_date}

@app.post("/api/transfers/save")
async def save_transfers_snapshot(payload: dict):
    import json as _j
    rd = payload.get("report_date")
    if not rd: raise HTTPException(400, "report_date required")
    conn = get_db()
    try:
        conn.execute("INSERT OR REPLACE INTO transfers_snapshots (report_date,saved_at,warehouses,data) VALUES(?,?,?,?)",
            (rd, datetime.now().isoformat()[:19],
             _j.dumps(payload.get("warehouses",[]), ensure_ascii=False),
             _j.dumps(payload.get("data",{}), ensure_ascii=False, separators=(",",":"))));
        conn.commit(); return {"ok": True, "report_date": rd}
    finally: conn.close()

@app.get("/api/transfers/list")
def list_transfers():
    import json as _j
    conn = get_db()
    try:
        rows = conn.execute("SELECT report_date,saved_at,warehouses FROM transfers_snapshots ORDER BY report_date DESC").fetchall()
        return [{"report_date":r[0],"saved_at":r[1],"warehouses":_j.loads(r[2])} for r in rows]
    finally: conn.close()

@app.get("/api/transfers/{report_date}")
def get_transfers(report_date: str):
    import json as _j
    conn = get_db()
    try:
        row = conn.execute("SELECT report_date,saved_at,warehouses,data FROM transfers_snapshots WHERE report_date=?",(report_date,)).fetchone()
        if not row: raise HTTPException(404,"Not found")
        return {"report_date":row[0],"saved_at":row[1],"warehouses":_j.loads(row[2]),"data":_j.loads(row[3])}
    finally: conn.close()

@app.delete("/api/transfers/{report_date}")
def del_transfers(report_date: str):
    conn = get_db()
    try:
        conn.execute("DELETE FROM transfers_snapshots WHERE report_date=?",(report_date,)); conn.commit(); return {"ok":True}
    finally: conn.close()


@app.delete("/api/stocks/date/{date_str}")
def delete_stocks_for_date(date_str: str):
    """Delete all stock records for a specific date so files can be re-uploaded."""
    conn = get_db()
    count = conn.execute("SELECT COUNT(*) as c FROM stock_snapshots WHERE date=?", (date_str,)).fetchone()["c"]
    conn.execute("DELETE FROM stock_snapshots WHERE date=?", (date_str,))
    conn.commit(); conn.close()
    # Invalidate caches
    global _analytics_cache, _analytics_cache_key
    _analytics_cache = None; _analytics_cache_key = None
    import os
    if os.path.exists(ANALYTICS_JSON_PATH): os.remove(ANALYTICS_JSON_PATH)
    if os.path.exists(TURNOVER_JSON_PATH): os.remove(TURNOVER_JSON_PATH)
    return {"ok": True, "deleted": count, "date": date_str}

@app.post("/api/admin/clear-sales")
def clear_sales_data():
    """Delete all sales_data rows so CSVs can be re-uploaded cleanly."""
    conn = get_db()
    old = conn.execute("SELECT COUNT(*) as c FROM sales_data WHERE row_id = ''").fetchone()["c"]
    all_ = conn.execute("SELECT COUNT(*) as c FROM sales_data").fetchone()["c"]
    conn.execute("DELETE FROM sales_data")
    conn.commit(); conn.close()
    return {"deleted": all_, "old_schema_rows": old, "message": "Залейте CSV заново."}

# ===================== ЯНДЕКС МАРКЕТ =====================
# Данные из финансового отчёта «По заказам» (united_orders_*.xlsx)
# Листы: «Услуги и маржа по заказам» (заказ + все услуги Маркета),
#        «Транзакции по заказам и товарам» (потоварные строки, 1 строка = 1 шт)
import json as _json

def _init_yandex():
    conn = get_db()
    conn.execute("""CREATE TABLE IF NOT EXISTS ym_orders (
        order_id TEXT PRIMARY KEY,
        date TEXT,
        status TEXT,
        price REAL DEFAULT 0,
        services REAL DEFAULT 0,
        svc_json TEXT DEFAULT '{}')""")
    conn.execute("""CREATE TABLE IF NOT EXISTS ym_items (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        order_id TEXT,
        date TEXT,
        sku TEXT,
        name TEXT,
        price REAL DEFAULT 0,
        status TEXT)""")
    conn.execute("""CREATE TABLE IF NOT EXISTS ym_money (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        order_id TEXT,
        kind TEXT,
        amount REAL DEFAULT 0,
        pp_date TEXT)""")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_ym_money_date ON ym_money(pp_date)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_ym_money_order ON ym_money(order_id)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_ym_items_order ON ym_items(order_id)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_ym_orders_date ON ym_orders(date)")
    conn.commit(); conn.close()

_init_yandex()

YM_SERVICES = [
    "Размещение товаров на витрине",
    "Складская обработка",
    "Программа лояльности и отзывы",
    "Буст продаж",
    "Рассрочка",
    "Доставка покупателю",
    "Доставка (средняя миля)",
    "Экспресс-доставка покупателю",
    "Доставка из-за рубежа",
    "Приём платежа покупателя",
    "Перевод платежа покупателя",
    "Организация забора заказов",
    "Обработка заказов в СЦ или ПВЗ",
    "Вывоз со склада, СЦ, ПВЗ",
    "Хранение невыкупов и возвратов",
    "Обработка заказов на складе",
    "Вознаграждение за продажу товара",
]

def _ym_date(v):
    s = str(v or "").strip()
    m = _re.match(r"(\d{2})\.(\d{2})\.(\d{4})", s)
    if m:
        return f"{m.group(3)}-{m.group(2)}-{m.group(1)}"
    m = _re.match(r"(\d{4})-(\d{2})-(\d{2})", s)
    if m:
        return s[:10]
    return None

def _ym_num(v):
    if v is None or v == "":
        return 0.0
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0

_YM_SIZE_RE = _re.compile(r"[\s,]+(?:XS|S|M|L|XL|XXL|XXXL|2XL|3XL|4XL|One\s?Size|ONE\s?SIZE|OS|\d{2}(?:-\d{2})?|\d{3}-\d{3})\s*$", _re.I)
def _ym_base(n):
    """Яндекс-название -> каноническое имя: убираем бренд, размеры и рост в конце."""
    n = _re.sub(r"\s*Chernim\s*Cherno\s*", " ", str(n or ""), flags=_re.I)
    prev = None
    while prev != n:
        prev = n
        n = _re.sub(r"\s*\([^)]*\)\s*$", "", n).strip()
        n = _YM_SIZE_RE.sub("", n).strip().rstrip(",")
    n = _re.sub(r"\s{2,}", " ", n).strip()
    return _canon_name(n) if n else n

def _ym_cls(status):
    s = str(status or "")
    if s.startswith("Доставлен"):
        return "delivered"
    if "Невыкуп" in s:
        return "nonbuyout"
    if "Возврат" in s:
        return "return"
    if s in ("Отменён", "Отменен", "Удалён", "Удален"):
        return "cancel"
    return "transit"

def _ym_find_header(ws, first_col_value, max_scan=20):
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=max_scan, values_only=True), start=1):
        if row and str(row[0] or "").strip() == first_col_value:
            return i, [str(c or "").strip() for c in row]
    return None, None

@app.post("/api/yandex/upload")
async def yandex_upload(file: UploadFile = File(...)):
    import openpyxl, warnings as _w
    _w.filterwarnings("ignore")
    raw = await file.read()
    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    tmp.write(raw); tmp.close()
    try:
        wb = openpyxl.load_workbook(tmp.name, data_only=True)
    except Exception as e:
        os.unlink(tmp.name)
        return {"error": f"Не удалось открыть файл: {e}"}
    need = ["Услуги и маржа по заказам", "Транзакции по заказам и товарам"]
    if any(s not in wb.sheetnames for s in need):
        os.unlink(tmp.name)
        return {"error": "Это не отчёт «По заказам» из кабинета Маркета (нет нужных листов)"}

    # ---- Лист услуг: заказ → все услуги ----
    ws = wb["Услуги и маржа по заказам"]
    hrow, hdr = _ym_find_header(ws, "ID бизнес-аккаунта")
    if not hrow:
        os.unlink(tmp.name)
        return {"error": "Не найдена шапка листа «Услуги и маржа по заказам»"}
    def col(name_part):
        for i, h in enumerate(hdr):
            if h.startswith(name_part):
                return i
        return None
    c_order = col("Номер заказа"); c_status = col("Статус заказа")
    c_date = col("Дата оформления"); c_total = col("Все услуги Маркета")
    c_price = col("Цена продажи")
    svc_cols = {}
    for s in YM_SERVICES:
        i = col(s)
        if i is not None:
            svc_cols[s] = i

    orders = {}
    for row in ws.iter_rows(min_row=hrow + 1, values_only=True):
        oid = row[c_order]
        if oid is None:
            continue
        oid = str(oid).split(".")[0]
        svc = {}
        for name, i in svc_cols.items():
            v = _ym_num(row[i])
            if v:
                svc[name] = round(v, 2)
        orders[oid] = {
            "date": _ym_date(row[c_date]),
            "status": str(row[c_status] or ""),
            "price": _ym_num(row[c_price]),
            "services": _ym_num(row[c_total]),
            "svc": svc,
        }

    # ---- Лист транзакций: потоварные строки ----
    ws2 = wb["Транзакции по заказам и товарам"]
    hrow2, hdr2 = _ym_find_header(ws2, "ID бизнес-аккаунта")
    if not hrow2:
        os.unlink(tmp.name)
        return {"error": "Не найдена шапка листа «Транзакции по заказам и товарам»"}
    def col2(name_part):
        for i, h in enumerate(hdr2):
            if h.startswith(name_part):
                return i
        return None
    t_order = col2("Номер заказа"); t_date = col2("Дата оформления")
    t_sku = col2("Ваш SKU"); t_name = col2("Название товара")
    t_price = col2("Цена продажи"); t_status = col2("Статус товара")

    # Группы платёжных колонок: (колонка суммы, колонка даты п/п, колонка даты реестра, тип)
    pay_groups = []
    grp_hdr = [str(c or "").strip() for c in next(ws2.iter_rows(min_row=hrow2 - 1, max_row=hrow2 - 1, values_only=True))]
    cur = ""
    sub = hdr2
    KIND = {"Платёж покупателя": "payment", "Платёж за скидку Маркета": "comp_market",
            "Платёж за скидку по бонусам СберСпасибо": "comp_sber", "Платёж за скидку Яндекс Плюс": "comp_plus",
            "Возврат платежа покупателя": "refund", "Возврат платежа за скидку Маркета": "refund_market",
            "Возврат платежа за скидку по бонусам СберСпасибо": "refund_sber",
            "Возврат платежа за скидку Яндекс Плюс": "refund_plus"}
    for i, g in enumerate(grp_hdr):
        if g:
            cur = g
        if sub[i].startswith("Сумма платежа") or sub[i].startswith("Сумма возврата") or sub[i].startswith("Удержанная сумма"):
            kind = KIND.get(cur, "withheld" if sub[i].startswith("Удержанная") else None)
            if kind:
                pay_groups.append((i, i + 2, i + 4, kind))

    items = []
    money = []
    for row in ws2.iter_rows(min_row=hrow2 + 2, values_only=True):
        oid = row[t_order]
        if oid is None:
            continue
        oid = str(oid).split(".")[0]
        name = str(row[t_name] or "").strip()
        if not name:
            continue
        d = _ym_date(row[t_date]) or (orders.get(oid) or {}).get("date")
        items.append((oid, d, str(row[t_sku] or "").strip(), name,
                      _ym_num(row[t_price]), str(row[t_status] or "")))
        for ci, di, ri, kind in pay_groups:
            v = _ym_num(row[ci])
            if v:
                pd = _ym_date(row[di]) or _ym_date(row[ri])
                if pd:
                    money.append((oid, kind, v, pd))
    os.unlink(tmp.name)

    if not orders or not items:
        return {"error": "В отчёте не нашлось данных"}

    conn = get_db()
    oids = list(orders.keys())
    CH = 500
    for i in range(0, len(oids), CH):
        chunk = oids[i:i + CH]
        ph = ",".join("?" * len(chunk))
        conn.execute(f"DELETE FROM ym_orders WHERE order_id IN ({ph})", chunk)
        conn.execute(f"DELETE FROM ym_items WHERE order_id IN ({ph})", chunk)
    conn.executemany(
        "INSERT OR REPLACE INTO ym_orders (order_id, date, status, price, services, svc_json) VALUES (?,?,?,?,?,?)",
        [(o, d["date"], d["status"], d["price"], d["services"], _json.dumps(d["svc"], ensure_ascii=False))
         for o, d in orders.items()])
    conn.executemany(
        "INSERT INTO ym_items (order_id, date, sku, name, price, status) VALUES (?,?,?,?,?,?)",
        items)
    for i in range(0, len(oids), CH):
        chunk = oids[i:i + CH]
        ph = ",".join("?" * len(chunk))
        conn.execute(f"DELETE FROM ym_money WHERE order_id IN ({ph})", chunk)
    conn.executemany(
        "INSERT INTO ym_money (order_id, kind, amount, pp_date) VALUES (?,?,?,?)",
        money)
    conn.commit()
    dates = sorted([d["date"] for d in orders.values() if d["date"]])
    conn.close()
    return {"ok": True, "orders": len(orders), "items": len(items),
            "date_from": dates[0] if dates else None, "date_to": dates[-1] if dates else None}

@app.get("/api/yandex/summary")
def yandex_summary(month: Optional[str] = None):
    conn = get_db()
    months = [r[0] for r in conn.execute(
        "SELECT DISTINCT substr(date,1,7) m FROM ym_orders WHERE date IS NOT NULL ORDER BY m DESC").fetchall()]
    if not months:
        conn.close()
        return {"months": [], "month": None, "summary": None, "services": [], "items": []}
    sel_all = (month == "all")
    if not sel_all and (not month or month not in months):
        month = months[0]

    if sel_all:
        orows = conn.execute("SELECT * FROM ym_orders").fetchall()
        irows = conn.execute("SELECT * FROM ym_items").fetchall()
    else:
        orows = conn.execute("SELECT * FROM ym_orders WHERE substr(date,1,7)=?", (month,)).fetchall()
        irows = conn.execute(
            "SELECT i.* FROM ym_items i JOIN ym_orders o ON o.order_id=i.order_id WHERE substr(o.date,1,7)=?",
            (month,)).fetchall()
    costs = {r["sku_base"]: r["cost"] for r in conn.execute("SELECT sku_base, cost FROM sku_costs").fetchall()}
    conn.close()

    # Аллокация услуг заказа на товары пропорционально цене
    order_items = {}
    for it in irows:
        order_items.setdefault(it["order_id"], []).append(it)

    svc_totals = {}
    total_services = 0.0
    for o in orows:
        total_services += o["services"] or 0
        for k, v in _json.loads(o["svc_json"] or "{}").items():
            svc_totals[k] = svc_totals.get(k, 0) + v

    skus = {}
    tot = {"delivered_qty": 0, "delivered_rev": 0.0, "nonbuyout_qty": 0, "nonbuyout_rev": 0.0,
           "return_qty": 0, "return_rev": 0.0, "cancel_qty": 0, "transit_qty": 0}

    omap = {o["order_id"]: o for o in orows}
    for oid, its in order_items.items():
        o = omap.get(oid)
        osvc = (o["services"] or 0) if o else 0
        psum = sum(max(i["price"] or 0, 0) for i in its)
        for it in its:
            share = ((it["price"] or 0) / psum) if psum > 0 else (1.0 / len(its))
            alloc = osvc * share
            base = _ym_base(it["name"])
            s = skus.setdefault(base, {
                "base": base, "delivered_qty": 0, "delivered_rev": 0.0,
                "nonbuyout_qty": 0, "return_qty": 0, "cancel_qty": 0, "transit_qty": 0,
                "services": 0.0})
            s["services"] += alloc
            cls = _ym_cls(it["status"])
            price = it["price"] or 0
            if cls == "delivered":
                s["delivered_qty"] += 1; s["delivered_rev"] += price
                tot["delivered_qty"] += 1; tot["delivered_rev"] += price
            elif cls == "nonbuyout":
                s["nonbuyout_qty"] += 1
                tot["nonbuyout_qty"] += 1; tot["nonbuyout_rev"] += price
            elif cls == "return":
                s["return_qty"] += 1
                tot["return_qty"] += 1; tot["return_rev"] += price
            elif cls == "cancel":
                s["cancel_qty"] += 1; tot["cancel_qty"] += 1
            else:
                s["transit_qty"] += 1; tot["transit_qty"] += 1

    items_out = []
    total_cost = 0.0
    no_cost = 0
    for base, s in skus.items():
        cost = costs.get(base) or 0
        cost_total = cost * s["delivered_qty"]
        profit = s["delivered_rev"] - s["services"] - cost_total
        attempted = s["delivered_qty"] + s["nonbuyout_qty"] + s["return_qty"]
        buyout = (s["delivered_qty"] / attempted * 100) if attempted else None
        if s["delivered_qty"] > 0:
            total_cost += cost_total
            if not cost:
                no_cost += 1
        items_out.append({
            **{k: (round(v, 2) if isinstance(v, float) else v) for k, v in s.items()},
            "cost": cost, "cost_total": round(cost_total, 2),
            "profit": round(profit, 2),
            "profit_unit": round(profit / s["delivered_qty"], 2) if s["delivered_qty"] else None,
            "buyout": round(buyout, 1) if buyout is not None else None,
            "flag": ("loss" if (profit < 0 or (cost and profit <= 0 and s["delivered_qty"] > 0))
                     else ("nocost" if (not cost and s["delivered_qty"] > 0) else "ok")),
        })
    items_out.sort(key=lambda x: x["profit"], reverse=True)

    ordered_rev = tot["delivered_rev"] + tot["nonbuyout_rev"] + tot["return_rev"]
    ordered_qty = tot["delivered_qty"] + tot["nonbuyout_qty"] + tot["return_qty"] + tot["cancel_qty"] + tot["transit_qty"]
    income = tot["delivered_rev"] - total_services
    summary = {
        **{k: (round(v, 2) if isinstance(v, float) else v) for k, v in tot.items()},
        "orders": len(orows),
        "ordered_rev": round(ordered_rev, 2), "ordered_qty": ordered_qty,
        "services_total": round(total_services, 2),
        "services_pct": round(total_services / tot["delivered_rev"] * 100, 1) if tot["delivered_rev"] else None,
        "income": round(income, 2),
        "cost_total": round(total_cost, 2),
        "profit": round(income - total_cost, 2),
        "no_cost_skus": no_cost,
        "buyout_rate": round(tot["delivered_qty"] /
                             (tot["delivered_qty"] + tot["nonbuyout_qty"] + tot["return_qty"]) * 100, 1)
                       if (tot["delivered_qty"] + tot["nonbuyout_qty"] + tot["return_qty"]) else None,
    }
    services = sorted([{"name": k, "sum": round(v, 2)} for k, v in svc_totals.items()],
                      key=lambda x: -x["sum"])
    return {"months": months, "month": ("all" if sel_all else month),
            "summary": summary, "services": services, "items": items_out}

@app.get("/api/yandex/money")
def yandex_money(month: Optional[str] = None):
    """Реальные деньги по датам платёжных поручений (возвраты в данных отрицательные)."""
    conn = get_db()
    months = [r[0] for r in conn.execute(
        "SELECT DISTINCT substr(pp_date,1,7) m FROM ym_money ORDER BY m DESC").fetchall()]
    if not months:
        conn.close()
        return {"months": [], "month": None, "money": None}
    sel_all = (month == "all")
    if not sel_all and (not month or month not in months):
        month = months[0]
    if sel_all:
        rows = conn.execute("SELECT kind, SUM(amount) s FROM ym_money GROUP BY kind").fetchall()
        svc = conn.execute("SELECT SUM(services) s FROM ym_orders").fetchone()["s"] or 0
    else:
        rows = conn.execute(
            "SELECT kind, SUM(amount) s FROM ym_money WHERE substr(pp_date,1,7)=? GROUP BY kind",
            (month,)).fetchall()
        svc = conn.execute(
            "SELECT SUM(services) s FROM ym_orders WHERE substr(date,1,7)=?", (month,)).fetchone()["s"] or 0
    conn.close()
    k = {r["kind"]: round(r["s"], 2) for r in rows}
    payments = k.get("payment", 0)
    comps = k.get("comp_market", 0) + k.get("comp_sber", 0) + k.get("comp_plus", 0)
    refunds = k.get("refund", 0) + k.get("refund_market", 0) + k.get("refund_sber", 0) + k.get("refund_plus", 0)
    withheld = k.get("withheld", 0)
    gross = payments + comps + refunds + withheld  # возвраты/удержания отрицательные в данных
    return {"months": months, "month": ("all" if sel_all else month),
            "money": {"payments": payments, "compensations": round(comps, 2),
                      "refunds": round(refunds, 2), "withheld": round(withheld, 2),
                      "gross": round(gross, 2),
                      "services_hint": round(svc, 2),
                      "net_estimate": round(gross - svc, 2)}}

@app.get("/yandex")
def serve_yandex():
    if os.path.exists("yandex.html"):
        return FileResponse("yandex.html", media_type="text/html")
    return FileResponse("index.html", media_type="text/html")


@app.get("/{full_path:path}")
def serve_frontend(full_path: str):
    if os.path.exists("index.html"):
        return FileResponse("index.html", media_type="text/html")
    return {"error": "not found"}
